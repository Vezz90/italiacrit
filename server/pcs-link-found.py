"""
Legge il file Excel compilato manualmente (colonna E "Link profilo trovato")
prodotto da atleti_non_trovati_pcs.xlsx, fa match del nome sull'anagrafica FCI
(data/athletes.json) ed emette pcs-found-links.json: [{atleta_id, name, slug, url}]
consumato poi da: node pcs-athlete-import.js --ids-file=pcs-found-links.json

Uso: python3 pcs-link-found.py <percorso_xlsx>
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

def normalize(s):
    return re.sub(r"[^A-Z ]", "", (s or "").upper()).strip()

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 pcs-link-found.py <percorso_xlsx>")
        sys.exit(1)
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"File non trovato: {xlsx_path}")
        sys.exit(1)

    athletes = json.loads((ROOT / "data" / "athletes.json").read_text(encoding="utf-8"))
    by_name = {}
    for aid, a in athletes.items():
        key = normalize(f"{a['cognome']} {a['nome']}")
        by_name.setdefault(key, []).append(aid)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    found = []
    unmatched = []
    ambiguous = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        name_cell, link_cell = row[1], row[4]  # colonne B (Nome) ed E (Link profilo trovato) — layout v7: #, Nome, Team, Cerca su PCS, Link, Note (niente colonna "Slug provati")
        name = (name_cell.value or "").strip()
        url  = (link_cell.value or "").strip() if link_cell.value else ""
        if not name or not url:
            continue
        if "procyclingstats.com/rider/" not in url:
            unmatched.append({"name": name, "url": url, "reason": "URL non è un profilo PCS rider/"})
            continue

        slug = url.rstrip("/").split("/rider/")[-1].split("/")[0]
        key = normalize(name)
        candidates = by_name.get(key, [])
        if len(candidates) == 1:
            found.append({"atleta_id": candidates[0], "name": name, "slug": slug, "url": url})
        elif len(candidates) > 1:
            ambiguous.append({"name": name, "url": url, "candidates": candidates})
        else:
            unmatched.append({"name": name, "url": url, "reason": "nome non trovato in athletes.json"})

    out_path = Path(__file__).resolve().parent / "pcs-found-links.json"
    out_path.write_text(json.dumps(found, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Abbinati:     {len(found)}  → {out_path}")
    print(f"Ambigui:      {len(ambiguous)} (stesso nome, più atleti — servono a mano)")
    for a in ambiguous:
        print(f"  - {a['name']}: {a['candidates']}")
    print(f"Non abbinati: {len(unmatched)}")
    for u in unmatched:
        print(f"  - {u['name']}: {u['reason']}")

if __name__ == "__main__":
    main()
